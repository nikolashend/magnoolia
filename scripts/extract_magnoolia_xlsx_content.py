"""
Phase 28 — Excel Content Extractor
Parses material/tile/finish Excel files from Magnoolia.
"""
import os
import sys
import json
from pathlib import Path
from datetime import datetime

ROOT = Path(__file__).parent.parent
IMPORT_DIR = ROOT / "materials" / "onedrive-import" / "phase28"
MATERIALS_DIR = ROOT / "materials"
OUT_DIR = ROOT / "storage" / "app" / "magnoolia" / "phase28" / "excel-extracted"
REPORT_PATH = ROOT / "docs" / "magnoolia-phase28-excel-content-report.md"

OUT_DIR.mkdir(parents=True, exist_ok=True)

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("ERROR: openpyxl not available")
    sys.exit(1)

FILES = [
    IMPORT_DIR / "Hals.xlsx",
    IMPORT_DIR / "Plaadid maht.xlsx",
    IMPORT_DIR / "Copy of Mag. tee ker plaadid.xlsx",
    MATERIALS_DIR / "_Magnoolia hinnatabel 05.05.26.xlsx",
]

def extract_workbook(path):
    result = {'file': path.name, 'path': str(path), 'sheets': []}
    if not path.exists():
        result['status'] = 'NOT_FOUND'
        return result

    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        result['status'] = 'OK'
        result['sheet_names'] = wb.sheetnames

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_data = []
            non_empty_rows = 0

            for row in ws.iter_rows(values_only=True):
                # Skip fully empty rows
                cells = [c for c in row if c is not None and str(c).strip() != '']
                if cells:
                    rows_data.append([str(c) if c is not None else '' for c in row])
                    non_empty_rows += 1

            sheet_info = {
                'name': sheet_name,
                'non_empty_rows': non_empty_rows,
                'rows': rows_data[:50],  # First 50 rows for inspection
            }
            result['sheets'].append(sheet_info)
            print(f"  Sheet '{sheet_name}': {non_empty_rows} non-empty rows")
            # Print first few rows
            for row in rows_data[:5]:
                print(f"    {' | '.join(str(c)[:40] for c in row[:6])}")

        wb.close()
    except Exception as e:
        result['status'] = f'ERROR: {e}'
        print(f"  Error: {e}")

    return result

def main():
    print("Phase 28 — Excel Content Extraction\n")
    all_results = []

    for f in FILES:
        print(f"\n{'='*60}")
        print(f"File: {f.name}")
        if not f.exists():
            print("  NOT FOUND — skipping")
            all_results.append({'file': f.name, 'status': 'NOT_FOUND'})
            continue
        size_kb = f.stat().st_size // 1024
        print(f"Size: {size_kb} KB")
        result = extract_workbook(f)
        all_results.append(result)

    # Save raw JSON
    out_json = OUT_DIR / "excel-content.json"
    with open(out_json, 'w', encoding='utf-8') as fp:
        json.dump(all_results, fp, ensure_ascii=False, indent=2)
    print(f"\nJSON saved: {out_json}")

    # Write markdown report
    lines = [
        "# Phase 28 Excel Content Report",
        "",
        f"Generated: {datetime.now().isoformat()}",
        "",
    ]

    for r in all_results:
        lines += [f"## {r['file']}", ""]
        if r.get('status') == 'NOT_FOUND':
            lines += ["**Status: NOT FOUND**", ""]
            continue
        if r.get('status', '').startswith('ERROR'):
            lines += [f"**Status: {r['status']}**", ""]
            continue

        lines += [f"**Status: OK** — Sheets: {', '.join(r.get('sheet_names', []))}", ""]

        for sheet in r.get('sheets', []):
            lines += [
                f"### Sheet: {sheet['name']}",
                f"- Non-empty rows: {sheet['non_empty_rows']}",
                "",
                "**First rows (preview):**",
                "",
                "```",
            ]
            for row in sheet['rows'][:20]:
                clean = [c[:50] for c in row if c.strip()]
                if clean:
                    lines.append(' | '.join(clean))
            lines += ["```", ""]

    REPORT_PATH.write_text('\n'.join(lines), encoding='utf-8')
    print(f"Report: {REPORT_PATH}")

    found = sum(1 for r in all_results if r.get('status') == 'OK')
    print(f"\nResult: {found}/{len(FILES)} files parsed successfully\n")

if __name__ == '__main__':
    main()
